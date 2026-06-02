"""
Excel de planificación ultra-compacto: todos los días en una sola hoja,
expandidos a la derecha. Cada día tiene su propio bloque (hora + campos),
separado por una columna vacía. El día se indica en la fila 1.
"""
import io
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

SLOT_MIN    = 5      # granularidad en minutos
EMPTY_H     = 2.0    # altura (pt) de filas vacías
MATCH_MIN_H = 16     # altura mínima total (pt) de un partido (1 línea)
COL_W       = 13     # ancho columnas de campo (chars)
HORA_COL_W  = 5.5    # ancho columna de hora
SEP_COL_W   = 2.0    # ancho columna separadora entre días

GRAY_EMPTY  = "FFCCCCCC"
OUTER_CLR   = "FF222222"
OUTER_SIDE  = Side(style="medium", color=OUTER_CLR)
HORA_FILL   = PatternFill("solid", fgColor="FFE0E0E0")
HDR_FILL    = PatternFill("solid", fgColor="FF1A3A5F")
HDR_FONT    = Font(bold=True, color="FFFFFFFF", size=8)
DAY_FILL    = PatternFill("solid", fgColor="FF0F2740")  # azul más oscuro para el día

_DIA_NAMES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
_MES_NAMES = ["ene", "feb", "mar", "abr", "may", "jun",
              "jul", "ago", "sep", "oct", "nov", "dic"]


def _hex_to_argb(css_hex: str) -> str:
    return "FF" + css_hex.lstrip("#").upper()


def _hm_to_min(t) -> int:
    try:
        parts = str(t)[:5].split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return 0


def _fecha_label(fecha_iso: str) -> str:
    try:
        dt = datetime.strptime(fecha_iso, "%Y-%m-%d")
        return f"{_DIA_NAMES[dt.weekday()]} {dt.day} {_MES_NAMES[dt.month - 1]}"
    except Exception:
        return fecha_iso


def _asignar_numeros_partido(partidos: list) -> dict:
    por_grupo: dict[str, list] = defaultdict(list)
    for p in partidos:
        por_grupo[p.get("grupo_id") or ""].append(p)
    nums: dict = {}
    for grupo_ps in por_grupo.values():
        ordenados = sorted(
            grupo_ps,
            key=lambda p: (p.get("jornada") or 0, str(p.get("hora") or ""), p.get("id") or "")
        )
        for i, p in enumerate(ordenados, start=1):
            nums[p.get("id")] = i
    return nums


def _add_border_side(cell, side_name: str, side: Side) -> None:
    b = cell.border
    kwargs = {"left": b.left, "right": b.right, "top": b.top, "bottom": b.bottom}
    kwargs[side_name] = side
    cell.border = Border(**kwargs)


def _apply_outer_border(ws, first_row: int, last_row: int,
                        first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        _add_border_side(ws.cell(row=first_row, column=col), "top",    OUTER_SIDE)
        _add_border_side(ws.cell(row=last_row,  column=col), "bottom", OUTER_SIDE)
    for row in range(first_row, last_row + 1):
        _add_border_side(ws.cell(row=row, column=first_col), "left",  OUTER_SIDE)
        _add_border_side(ws.cell(row=row, column=last_col),  "right", OUTER_SIDE)


def _set_merged_borders(ws, start_row: int, end_row: int, col: int, bside: Side) -> None:
    for r in range(start_row, end_row + 1):
        top_s = bside if r == start_row else Side(style=None)
        bot_s = bside if r == end_row   else Side(style=None)
        ws.cell(row=r, column=col).border = Border(
            left=bside, right=bside, top=top_s, bottom=bot_s
        )


def generar_excel_planificacion(
    partidos: list,
    titulo: str,
    hora_ini_min: int = 480,
    hora_fin_min: int = 1320,
    color_map: dict | None = None,
) -> bytes:
    if hora_fin_min <= hora_ini_min:
        hora_fin_min = hora_ini_min + 60

    color_map    = color_map or {}
    nums_partido = _asignar_numeros_partido(partidos)

    torneo_names: dict[str, str] = {}
    for p in partidos:
        tid = p.get("torneo_id") or ""
        if tid and tid not in torneo_names:
            torneo_names[tid] = p.get("nombre_torneo") or tid

    por_fecha: dict = defaultdict(lambda: defaultdict(list))
    for p in partidos:
        fecha = str(p.get("fecha") or "")[:10]
        campo = p.get("campo") or "Sin campo"
        por_fecha[fecha][campo].append(p)

    # ── Rango horario global (mismo para todos los días) ──────────────────────
    all_starts = [hora_ini_min] if hora_ini_min else []
    all_ends   = [hora_fin_min] if hora_fin_min else []
    for campos_dia in por_fecha.values():
        for campo_ps in campos_dia.values():
            for p in campo_ps:
                m   = _hm_to_min(str(p.get("hora") or "")[:5])
                dur = p.get("duracion_partido") or 50
                if m > 0:
                    all_starts.append(m)
                    all_ends.append(m + dur)

    g_ini = (min(all_starts) // 30) * 30 if all_starts else (hora_ini_min or 480)
    g_fin = ((max(all_ends)  + 29) // 30) * 30 if all_ends else (hora_fin_min or 1320)
    if hora_ini_min:
        g_ini = min(g_ini, hora_ini_min)
    if hora_fin_min:
        g_fin = max(g_fin, hora_fin_min)
    if g_fin <= g_ini:
        g_fin = g_ini + 60

    slots   = list(range(g_ini, g_fin, SLOT_MIN))
    n_slots = len(slots)

    # Filas: 1=día, 2=cabecera campos, 3..n_slots+2=datos
    ROW_DAY   = 1
    ROW_HDR   = 2
    ROW_DATA  = 3
    last_data_row = ROW_DATA + n_slots - 1

    # ── Preparar libro ────────────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title           = (titulo or "Planificación")[:31]
    ws.freeze_panes    = "A3"   # las dos filas de cabecera quedan fijas

    ws.row_dimensions[ROW_DAY].height = 11
    ws.row_dimensions[ROW_HDR].height = 13
    for idx in range(n_slots):
        ws.row_dimensions[ROW_DATA + idx].height = EMPTY_H

    # Dict global de alturas de fila necesarias por partidos
    row_heights: dict[int, float] = {}
    # Celdas de campo ocupadas (col, row) — global para todas las columnas
    occupied: set[tuple[int, int]] = set()

    # ── Bloque horario para la columna de horas (igual en todos los días) ─────
    hour_blocks: list[tuple[int, int, str]] = []
    block_start = 0
    block_label = f"{slots[0] // 60:02d}:{slots[0] % 60:02d}"
    for idx, slot_min in enumerate(slots):
        if slot_min % 60 == 0 and idx > 0:
            hour_blocks.append((block_start, idx - 1, block_label))
            block_start = idx
            block_label = f"{slot_min // 60:02d}:00"
    hour_blocks.append((block_start, len(slots) - 1, block_label))

    # ── Procesar cada día ─────────────────────────────────────────────────────
    col_offset = 0   # desplazamiento de columna (0-indexed)

    for fecha_iso in sorted(por_fecha.keys()):
        campos_dia = por_fecha[fecha_iso]
        campo_list = sorted(campos_dia.keys())
        n_campos   = len(campo_list)

        hora_col        = col_offset + 1           # columna hora (1-indexed)
        first_campo_col = col_offset + 2
        last_col        = col_offset + 1 + n_campos

        # Anchos de columna
        ws.column_dimensions[get_column_letter(hora_col)].width = HORA_COL_W
        for i in range(n_campos):
            ws.column_dimensions[get_column_letter(first_campo_col + i)].width = COL_W
        # Columna separadora (después de last_col)
        ws.column_dimensions[get_column_letter(last_col + 1)].width = SEP_COL_W

        # ── Fila 1: nombre del día ────────────────────────────────────────────
        # Celda de hora en fila 1
        ws.cell(row=ROW_DAY, column=hora_col).fill = DAY_FILL

        # Nombre del día, fusionado sobre las columnas de campo
        if n_campos > 1:
            ws.merge_cells(start_row=ROW_DAY, start_column=first_campo_col,
                           end_row=ROW_DAY,   end_column=last_col)
        day_cell = ws.cell(row=ROW_DAY, column=first_campo_col,
                           value=_fecha_label(fecha_iso).upper())
        day_cell.fill      = DAY_FILL
        day_cell.font      = Font(bold=True, color="FFFFFFFF", size=8)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Fila 2: cabecera de campos ────────────────────────────────────────
        ws.cell(row=ROW_HDR, column=hora_col).fill = HDR_FILL
        ws.cell(row=ROW_HDR, column=hora_col).font = HDR_FONT

        for i, campo in enumerate(campo_list):
            lbl = campo.upper()
            max_chars = int(COL_W * 1.3)
            if len(lbl) > max_chars:
                lbl = lbl[:max_chars - 1] + "…"
            cell = ws.cell(row=ROW_HDR, column=first_campo_col + i, value=lbl)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Partidos ──────────────────────────────────────────────────────────
        for campo in campo_list:
            col = first_campo_col + campo_list.index(campo)
            ps  = sorted(campos_dia[campo], key=lambda p: str(p.get("hora") or ""))

            for p in ps:
                hora_str = str(p.get("hora") or "")[:5]
                if not hora_str:
                    continue
                p_min   = _hm_to_min(hora_str)
                dur     = p.get("duracion_partido") or 50
                end_min = p_min + dur

                if p_min >= g_fin or end_min <= g_ini:
                    continue

                start_idx = max(0, (p_min - g_ini) // SLOT_MIN)
                n_rows    = max(1, -(-dur // SLOT_MIN))
                end_idx   = min(start_idx + n_rows - 1, n_slots - 1)
                start_row = start_idx + ROW_DATA
                end_row   = end_idx   + ROW_DATA

                if any((col, r) in occupied for r in range(start_row, end_row + 1)):
                    continue
                for r in range(start_row, end_row + 1):
                    occupied.add((col, r))

                tid = p.get("torneo_id") or ""
                bg_hex, border_hex = color_map.get(tid, ("#FFFFFF", "#D0D0D0"))
                fill        = PatternFill("solid", fgColor=_hex_to_argb(bg_hex))
                border_side = Side(style="medium", color=_hex_to_argb(border_hex))

                grp     = p.get("nombre_grupo") or "—"
                n       = nums_partido.get(p.get("id"), 1)
                content = f"{grp} - P{n}"

                per_row_h = MATCH_MIN_H / (end_row - start_row + 1)
                for r in range(start_row, end_row + 1):
                    row_heights[r] = max(row_heights.get(r, EMPTY_H), per_row_h)

                if start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=col,
                                   end_row=end_row, end_column=col)

                cell = ws.cell(row=start_row, column=col, value=content)
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=False)
                cell.font      = Font(size=7, bold=True, color=_hex_to_argb(border_hex))
                _set_merged_borders(ws, start_row, end_row, col, border_side)

        # ── Gris en celdas de campo vacías ────────────────────────────────────
        gray_fill = PatternFill("solid", fgColor=GRAY_EMPTY)
        inner_brd = Border(left=Side(style="thin", color="FFBBBBBB"),
                           right=Side(style="thin", color="FFBBBBBB"))
        for idx in range(n_slots):
            row = ROW_DATA + idx
            for i in range(n_campos):
                col = first_campo_col + i
                if (col, row) not in occupied:
                    c = ws.cell(row=row, column=col)
                    c.fill   = gray_fill
                    c.border = inner_brd

        # ── Columna de horas: combinar por bloque horario ─────────────────────
        for (s_idx, e_idx, label) in hour_blocks:
            s_row = s_idx + ROW_DATA
            e_row = e_idx + ROW_DATA
            if s_row < e_row:
                ws.merge_cells(start_row=s_row, start_column=hora_col,
                               end_row=e_row,   end_column=hora_col)
            cell = ws.cell(row=s_row, column=hora_col, value=label)
            cell.fill      = HORA_FILL
            cell.font      = Font(size=6, bold=True, color="FF333333")
            cell.alignment = Alignment(horizontal="right", vertical="top")

        # ── Borde exterior grueso ─────────────────────────────────────────────
        _apply_outer_border(ws, first_row=ROW_DAY, last_row=last_data_row,
                            first_col=hora_col, last_col=last_col)

        col_offset += 1 + n_campos + 1   # hora + campos + separador

    # ── Alturas finales de filas con partidos ─────────────────────────────────
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = max(h, EMPTY_H)

    # ── Leyenda debajo del primer bloque ─────────────────────────────────────
    all_torneos: dict[str, str] = {}
    for p in partidos:
        tid = p.get("torneo_id") or ""
        if tid and tid not in all_torneos:
            all_torneos[tid] = torneo_names.get(tid, tid)

    if all_torneos:
        ley_row0 = last_data_row + 2
        ws.row_dimensions[ley_row0].height = 13
        ley_title = ws.cell(row=ley_row0, column=1, value="Leyenda")
        ws.merge_cells(start_row=ley_row0, start_column=1,
                       end_row=ley_row0, end_column=3)
        ley_title.fill      = HDR_FILL
        ley_title.font      = Font(bold=True, color="FFFFFFFF", size=8)
        ley_title.alignment = Alignment(horizontal="left", vertical="center")

        for i, (tid, nombre) in enumerate(all_torneos.items()):
            row = ley_row0 + 1 + i
            ws.row_dimensions[row].height = 12
            bg_hex, border_hex = color_map.get(tid, ("#F1F5F9", "#334155"))
            fill  = PatternFill("solid", fgColor=_hex_to_argb(bg_hex))
            bside = Side(style="medium", color=_hex_to_argb(border_hex))
            bord  = Border(left=bside, right=bside, top=bside, bottom=bside)

            sw = ws.cell(row=row, column=1, value="")
            sw.fill = fill; sw.border = bord

            nc = ws.cell(row=row, column=2, value=nombre)
            nc.fill      = fill
            nc.border    = bord
            nc.font      = Font(size=8, bold=True, color=_hex_to_argb(border_hex))
            nc.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
