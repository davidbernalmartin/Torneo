"""Excel agenda: time grid, campos as columns, matches as merged colored cells."""
import io
import urllib.request
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
try:
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    _HAS_XL_IMG = True
except ImportError:
    _HAS_XL_IMG = False

RFFM_LOGO_URL = "https://rffm-cms.s3.eu-west-1.amazonaws.com/large_favicon_87ea61909c.png"
SLOT_MIN      = 5    # granularidad de la rejilla en minutos
_XL_SHIELD_PX = 22   # tamaño del escudo en pixels dentro de la celda
# Ancho aproximado de una columna de campo (29 chars) en pixels: 29*7+5 = 208 px
_COL_W_PX = 29 * 7 + 5


def _hex_to_argb(css_hex: str) -> str:
    return "FF" + css_hex.lstrip("#").upper()


def _hm_to_min(t) -> int:
    try:
        parts = str(t)[:5].split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return 0


def _fetch_bytes(url: str | None, cache: dict) -> bytes | None:
    """Descarga una imagen y devuelve los bytes en crudo, con caché."""
    if not url:
        return None
    if url in cache:
        return cache[url]
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=6) as r:
            data = r.read()
        cache[url] = data
        return data
    except Exception:
        cache[url] = None
        return None


def _add_xl_shield(ws, data: bytes | None, fallback: bytes | None,
                   col_idx: int, row_idx: int,
                   col_off_emu: int, row_off_emu: int) -> None:
    """Inserta un escudo en la hoja anclado en (col_idx, row_idx) con offsets en EMU."""
    if not _HAS_XL_IMG:
        return
    raw = data or fallback
    if not raw:
        return
    try:
        img = XLImage(io.BytesIO(raw))
        img.width  = _XL_SHIELD_PX
        img.height = _XL_SHIELD_PX
        sz = _XL_SHIELD_PX * 9525  # px → EMU
        marker = AnchorMarker(col=col_idx, colOff=col_off_emu,
                              row=row_idx, rowOff=row_off_emu)
        img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(sz, sz))
        ws.add_image(img)
    except Exception:
        pass


# Altura estimada de contenido: fuente 8pt → ~11pt por línea + 7pt de padding (+10 %)
_LINE_H = 11  # pt por línea de texto
_PAD_H  = 7   # pt de padding total (arriba + abajo)


def generar_excel_agenda(
    partidos: list,
    titulo: str,
    hora_ini_min: int = 480,
    hora_fin_min: int = 1320,
    color_map: dict | None = None,
) -> bytes:
    """Returns bytes of an .xlsx with one sheet per day, time grid + legend sheet."""
    if hora_fin_min <= hora_ini_min:
        hora_fin_min = hora_ini_min + 60

    color_map = color_map or {}

    # Extraer nombres de torneos de los partidos
    torneo_names: dict[str, str] = {}
    for p in partidos:
        tid = p.get("torneo_id") or ""
        if tid and tid not in torneo_names:
            torneo_names[tid] = p.get("nombre_torneo") or tid

    por_fecha: dict = defaultdict(lambda: defaultdict(list))
    for p in partidos:
        fecha = str(p.get("fecha") or "")[:10]
        campo = p.get("campo") or "Sin campo"
        por_fecha[fecha][campo].append(p)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    HEADER_FILL = PatternFill("solid", fgColor="FF1A3A5F")
    HEADER_FONT_W = Font(bold=True, color="FFFFFFFF", size=9)

    # Caché de imágenes para no descargar el mismo escudo varias veces
    _img_cache: dict = {}

    thin_grey = Side(style="thin", color="FFCCCCCC")
    DEFAULT_BORDER = Border(left=thin_grey, right=thin_grey, top=thin_grey, bottom=thin_grey)

    def _dia_ini(campos_dia: dict) -> int:
        """Inicio del primer partido del día, redondeado a 30 min y con suelo hora_ini_min."""
        ini = 23 * 60
        for campo_ps in campos_dia.values():
            for p in campo_ps:
                p_min = _hm_to_min(str(p.get("hora") or "")[:5])
                if p_min > 0:
                    ini = min(ini, p_min)
        if ini >= 23 * 60:
            ini = hora_ini_min or (8 * 60)
        ini = (ini // 30) * 30  # redondear hacia abajo a 30 min
        return min(ini, hora_ini_min) if hora_ini_min else ini  # si está fijado, usar el más temprano

    def _dia_fin(campos_dia: dict, ini_min: int) -> int:
        """Fin del último partido del día, redondeado a 30 min y capado por hora_fin_min."""
        if hora_fin_min:
            return hora_fin_min
        fin = ini_min + 60
        for campo_ps in campos_dia.values():
            for p in campo_ps:
                p_min = _hm_to_min(str(p.get("hora") or "")[:5])
                fin = max(fin, p_min + (p.get("duracion_partido") or 50))
        fin = (fin + 29) // 30 * 30
        return fin

    def _set_merged_borders(ws, start_row: int, end_row: int, col: int, bside: Side) -> None:
        """Aplica bordes correctamente a todas las celdas de un rango fusionado."""
        for r in range(start_row, end_row + 1):
            top_s   = bside if r == start_row else Side(style=None)
            bot_s   = bside if r == end_row   else Side(style=None)
            ws.cell(row=r, column=col).border = Border(
                left=bside, right=bside, top=top_s, bottom=bot_s
            )

    for fecha_iso in sorted(por_fecha.keys()):
        campos_dia = por_fecha[fecha_iso]
        campo_list = sorted(campos_dia.keys())

        hora_ini_dia = _dia_ini(campos_dia)
        hora_fin_dia = _dia_fin(campos_dia, hora_ini_dia)
        slots = list(range(hora_ini_dia, hora_fin_dia, SLOT_MIN))
        n_slots = len(slots)

        ws = wb.create_sheet(title=fecha_iso[:31])
        ws.freeze_panes = "B2"

        # Column widths
        ws.column_dimensions["A"].width = 7
        for i in range(len(campo_list)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 29

        # Header row
        ws.row_dimensions[1].height = 22
        cell_a1 = ws.cell(row=1, column=1, value="Hora")
        cell_a1.fill = HEADER_FILL
        cell_a1.font = HEADER_FONT_W
        cell_a1.alignment = Alignment(horizontal="center", vertical="center")

        for i, campo in enumerate(campo_list):
            cell = ws.cell(row=1, column=i + 2, value=campo.upper())
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT_W
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Time axis — altura pequeña para slots vacíos; etiqueta solo cada 30 min
        for idx, slot_min in enumerate(slots):
            row = idx + 2
            h, m = slot_min // 60, slot_min % 60
            show_label = (m % 30 == 0)
            ws.row_dimensions[row].height = 8 if not show_label else 10
            if show_label:
                cell = ws.cell(row=row, column=1, value=f"{h:02d}:{m:02d}")
                cell.font = Font(size=7, bold=True, color="FF222222")
                cell.alignment = Alignment(horizontal="right", vertical="center")
            fill_bg = "FFF0F0F0" if m == 0 else ("FFFAFAFA" if m == 30 else "FFFFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=fill_bg)

        # Track occupied (col, row) pairs, needed row heights e imágenes diferidas
        occupied: set[tuple[int, int]] = set()
        row_heights: dict[int, float] = {}       # row → max height needed (pt)
        deferred_imgs: list = []                 # (esc_l_bytes, esc_v_bytes, col, sr, er)

        for campo in campo_list:
            col = campo_list.index(campo) + 2
            ps = sorted(campos_dia[campo], key=lambda p: str(p.get("hora") or ""))

            for p in ps:
                hora_str = str(p.get("hora") or "")[:5]
                if not hora_str:
                    continue
                p_min = _hm_to_min(hora_str)
                dur = p.get("duracion_partido") or 50
                end_min = p_min + dur

                if p_min >= hora_fin_dia or end_min <= hora_ini_dia:
                    continue

                # Snap to slot grid
                start_idx = max(0, (p_min - hora_ini_dia) // SLOT_MIN)
                n_rows = max(1, -(-dur // SLOT_MIN))  # ceiling div
                end_idx = min(start_idx + n_rows - 1, n_slots - 1)

                start_row = start_idx + 2
                end_row = end_idx + 2

                if any((col, r) in occupied for r in range(start_row, end_row + 1)):
                    continue
                for r in range(start_row, end_row + 1):
                    occupied.add((col, r))

                # Colors — fondo pastel, texto y borde en el acento oscuro
                tid = p.get("torneo_id") or ""
                bg_hex, border_hex = color_map.get(tid, ("#FFFFFF", "#D0D0D0"))
                fill = PatternFill("solid", fgColor=_hex_to_argb(bg_hex))
                border_side = Side(style="medium", color=_hex_to_argb(border_hex))

                # Content
                loc = p.get("nombre_local") or "—"
                vis = p.get("nombre_visitante") or "—"
                rl, rv = p.get("resultado_local"), p.get("resultado_visitante")
                marc = f"  {rl}–{rv}" if rl is not None and rv is not None else ""
                torneo = p.get("nombre_torneo") or ""
                grp = p.get("nombre_grupo") or ""
                lines = [" · ".join(filter(None, [grp, torneo])), loc + marc, "vs", vis]
                content = "\n".join(filter(None, lines))

                # Calculate needed height and distribute across merged rows
                n_lines = len([ln for ln in lines if ln])
                total_content_h = n_lines * _LINE_H + _PAD_H
                per_row_h = total_content_h / (end_row - start_row + 1)
                for r in range(start_row, end_row + 1):
                    row_heights[r] = max(row_heights.get(r, 8), per_row_h)

                # Merge and write cell
                if start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=col,
                                   end_row=end_row, end_column=col)

                cell = ws.cell(row=start_row, column=col, value=content)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(size=8, bold=True, color=_hex_to_argb(border_hex))

                # Fix borders across all cells in the merged range
                _set_merged_borders(ws, start_row, end_row, col, border_side)

                # Guardar escudos para insertar después (necesitamos las alturas finales)
                esc_l = _fetch_bytes(p.get("escudo_local"), _img_cache)
                esc_v = _fetch_bytes(p.get("escudo_visitante"), _img_cache)
                deferred_imgs.append((esc_l, esc_v, col, start_row, end_row))

        # Apply computed row heights
        for r, h in row_heights.items():
            ws.row_dimensions[r].height = max(h, 8)

        # Insertar escudos ahora que las alturas definitivas están fijadas
        if _HAS_XL_IMG:
            sz_emu   = _XL_SHIELD_PX * 9525
            left_off = 4 * 9525                              # 4 px desde el borde izq
            right_off = max(0, (_COL_W_PX - _XL_SHIELD_PX - 4) * 9525)  # cerca del borde der
            for esc_l, esc_v, col, sr, er in deferred_imgs:
                total_h_pt  = sum(ws.row_dimensions[r].height for r in range(sr, er + 1))
                row_off_emu = max(0, int(total_h_pt * 12700 - sz_emu) // 2)
                _add_xl_shield(ws, esc_l, None, col - 1, sr - 1, left_off,  row_off_emu)
                _add_xl_shield(ws, esc_v, None, col - 1, sr - 1, right_off, row_off_emu)

        # Light background for empty cells
        for idx in range(n_slots):
            row = idx + 2
            for col_i in range(len(campo_list)):
                col = col_i + 2
                if (col, row) not in occupied:
                    c = ws.cell(row=row, column=col)
                    if not c.value:
                        bg = "FFF7F7F7" if (slots[idx] // 60) % 2 == 0 else "FFFFFFFF"
                        c.fill = PatternFill("solid", fgColor=bg)
                        c.border = DEFAULT_BORDER

        # ── Leyenda al pie de cada hoja ───────────────────────────────────────
        torneos_del_dia: dict[str, str] = {}
        for campo_ps in campos_dia.values():
            for p in campo_ps:
                tid = p.get("torneo_id") or ""
                if tid and tid not in torneos_del_dia:
                    torneos_del_dia[tid] = torneo_names.get(tid, tid)

        if torneos_del_dia:
            last_col = len(campo_list) + 1   # última columna del grid
            ley_row0 = n_slots + 3           # fila de cabecera de la leyenda

            # Título de leyenda solo en cols A y B
            ws.row_dimensions[ley_row0].height = 16
            ley_title = ws.cell(row=ley_row0, column=1, value="Leyenda")
            ws.merge_cells(start_row=ley_row0, start_column=1,
                           end_row=ley_row0, end_column=2)
            ley_title.fill = HEADER_FILL
            ley_title.font = HEADER_FONT_W
            ley_title.alignment = Alignment(horizontal="left", vertical="center")

            for i, (tid, nombre) in enumerate(torneos_del_dia.items()):
                row = ley_row0 + 1 + i
                ws.row_dimensions[row].height = 16
                bg_hex, border_hex = color_map.get(tid, ("#F1F5F9", "#334155"))
                fill  = PatternFill("solid", fgColor=_hex_to_argb(bg_hex))
                bside = Side(style="medium", color=_hex_to_argb(border_hex))
                bord  = Border(left=bside, right=bside, top=bside, bottom=bside)

                # Franja de color en col A (eje de horas)
                swatch = ws.cell(row=row, column=1, value="")
                swatch.fill = fill
                swatch.border = bord

                # Nombre del torneo solo en col B
                name_cell = ws.cell(row=row, column=2, value=nombre)
                name_cell.fill = fill
                name_cell.border = bord
                name_cell.font = Font(size=9, bold=True, color=_hex_to_argb(border_hex))
                name_cell.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
