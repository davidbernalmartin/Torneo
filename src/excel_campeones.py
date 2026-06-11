"""
Excel de campeones y subcampeones: una fila por torneo, tres columnas
(nombre torneo | campeón | subcampeón), sin colores ni escudos.
"""
import io

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="FF000000")
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def generar_excel_campeones(datos: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campeones"

    hdr_font  = Font(bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    data_font  = Font(size=10)

    headers = ["TORNEO", "CAMPEÓN", "SUBCAMPEÓN", "EQUIPOS", "PARTIDOS"]
    col_widths = [30, 30, 30, 10, 10]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = BRD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 18

    num_align = Alignment(horizontal="center", vertical="center")

    for ri, d in enumerate(datos, 2):
        ws.row_dimensions[ri].height = 16
        row_values = [
            d.get("torneo_nombre", ""),
            d.get("campeon",    {}).get("nombre", "") if d.get("campeon")    else "",
            d.get("subcampeon", {}).get("nombre", "") if d.get("subcampeon") else "",
            d.get("num_equipos",  0),
            d.get("num_partidos", 0),
        ]
        for ci, value in enumerate(row_values, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font      = data_font
            cell.alignment = num_align if ci >= 4 else data_align
            cell.border    = BRD

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
